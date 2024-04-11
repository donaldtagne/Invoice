
import tkinter
import os
import uuid
import tkinter.messagebox as mb
from tkinter import ttk
import tkinter as tk
from tkinter import PhotoImage
from PIL import Image, ImageTk
from tkcalendar import DateEntry
import pandas as pd
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4, portrait
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import Image, Table
from docxtpl import DocxTemplate
import datetime
import comtypes.client

class Myfacture(tkinter.Frame):
    
    def __init__(self, master=None):
        super().__init__(master)
        self.pack()
        self.color = "lightblue"
        self.menuBar = tkinter.Menu(master)
        master.config(menu=self.menuBar)
        self.Menubar()
        self.createwidgets()
        
        
    def Menubar(self):
      # Menü "Fichier"
        self.menuFile = tk.Menu(self.menuBar, tearoff=False)
        self.menuFile.add_command(label="Fermer", foreground="black", command=lambda: root.destroy())
        self.menuBar.add_cascade(label="Fichier", menu=self.menuFile)

        # Menü "Options"
        self.menu_opt = tk.Menu(self.menuBar, tearoff=False)
        self.menuBar.add_cascade(label="Options", menu=self.menu_opt)

        # Untermenü für "Numark"
        self.selected_numark = tk.StringVar()
        self.menu_numark = tk.Menu(self.menu_opt, tearoff=0)
        self.menu_numark.add_radiobutton(label='DJ Controller', variable=self.selected_numark, value="DJ Controller", command=self.update_combobox_values)
        self.menu_numark.add_radiobutton(label='Player', variable=self.selected_numark, value="Player", command=self.update_combobox_values)
        self.menu_numark.add_radiobutton(label='Mixers', variable=self.selected_numark, value="Mixers", command=self.update_combobox_values)
        self.menu_numark.add_radiobutton(label='Platines Vinyles', variable=self.selected_numark, value="Platines Vinyles", command=self.update_combobox_values)
        self.menu_numark.add_radiobutton(label='Cartouches', variable=self.selected_numark, value="Cartouches", command=self.update_combobox_values)
        self.menu_numark.add_radiobutton(label='Accessoires', variable=self.selected_numark, value="Accessoires", command=self.update_combobox_values)
        self.menu_opt.add_cascade(label="Numark", menu=self.menu_numark)

        # Untermenü für "Pioneer"
        self.menu_pioneer = tk.Menu(self.menu_opt, tearoff=0)
        self.selected_pioneer = tk.StringVar()
        self.menu_pioneer.add_radiobutton(label='Lecteur DJ/tourne-disque', variable=self.selected_pioneer, value="Lecteur DJ/tourne-disque", command=self.update_combobox_values)
        self.menu_pioneer.add_radiobutton(label='DJ Mixers', variable=self.selected_pioneer, value="DJ Mixers", command=self.update_combobox_values)
        self.menu_pioneer.add_radiobutton(label='DJ Controllern', variable=self.selected_pioneer, value="DJ Controllern", command=self.update_combobox_values)
        self.menu_pioneer.add_radiobutton(label='All-in-one-DJ-Systeme', variable=self.selected_pioneer, value="All-in-one-DJ-Systeme", command=self.update_combobox_values)
        self.menu_pioneer.add_radiobutton(label='Haut-Parleurs de Controle', variable=self.selected_pioneer, value="Haut-Parleurs de Controle", command=self.update_combobox_values)
        self.menu_pioneer.add_radiobutton(label='Appareil a Effets DJ', variable=self.selected_pioneer, value="Appareil a Effets DJ", command=self.update_combobox_values)
        self.menu_pioneer.add_radiobutton(label='Ecouteurs', variable=self.selected_pioneer, value="Ecouteurs", command=self.update_combobox_values)
        self.menu_pioneer.add_radiobutton(label='Accesoires', variable=self.selected_pioneer, value="Accesoires", command=self.update_combobox_values)
        self.menu_pioneer.add_radiobutton(label='Baffle', variable=self.selected_pioneer, value="Baffle", command=self.update_combobox_values)
        self.menu_opt.add_cascade(label="Pioneer", menu=self.menu_pioneer)

        self.vendeur = tkinter.Menu(self.menuBar, tearoff=False)
        self.selected_seller = tkinter.IntVar() 
        self.vendeur.add_radiobutton(label="Yannick", foreground='black', variable=self.selected_seller, value=1,  command=self.write_into)
        self.vendeur.add_separator()
        self.vendeur.add_radiobutton(label="Guyso", foreground='black', variable=self.selected_seller, value=2, command=self.write_into)
        self.vendeur.add_separator()
        self.vendeur.add_radiobutton(label="Merlin", foreground='black', variable=self.selected_seller, value=3, command=self.write_into)
        self.menuBar.add_cascade(label="Vendeur", menu=self.vendeur)
        
        self.menuHilfe = tkinter.Menu(self.menuBar, tearoff=False)
        self.menuHilfe.add_command(label="Contact",foreground='black', command= (lambda: self.openHelp()))
        self.menuBar.add_cascade(label="Soutient", menu=self.menuHilfe)
        
    def createwidgets(self): 
        self.tabControl = ttk.Notebook(root)
        
        # tab1
        self.tab1 = ttk.Frame(self.tabControl)
        self.tabControl.add(self.tab1, text ='Espace de Facture')
        self.tabControl.pack(expand = 1, fill ="both")
        # element und widgets controlle(vendeur)
        
        # premiere element
        self.frame = tkinter.LabelFrame(self.tab1, text="Vendeur", padx=20, pady=10)
        self.frame.grid(column=0, row=0, padx=10, pady=10, sticky="w")
        self.nom = tkinter.Label(self.frame, text="Nom de l'Entreprise")
        self.nom.grid(column=1, row=0, padx=5, sticky="w")
        self.nom1= tkinter.Entry(self.frame, width =30)
        self.nom1.grid(column =2, row=0)
        
        #gereur despace entre les elements
        self.frame.grid_rowconfigure(1, minsize=50)
        
        #deuxieme element
        self.adresse = tkinter.Label(self.frame, text="Adresse/Quatier")
        self.adresse.grid(column=1, row=1, padx=5, sticky="w")
        self.adresse2= tkinter.Entry(self.frame, width =30)
        self.adresse2.grid(column =2, row=1)
        #troisieme element
        self.contact = tkinter.Label(self.frame, text="Contact")
        self.contact.grid(column=1, row=2, padx=5, sticky="w")
        self.entre= tkinter.Entry(self.frame, width =30)
        self.entre.grid(column =2, row=2)
         
        # element und widgets controlle(achecteur)
        self.frame = tkinter.LabelFrame(self.tab1, text="Acheteur", padx=20, pady=10)
        self.frame.grid(column=1, row=0, padx=10, pady=10, sticky="w")
        self.frame.place(x=459, y=10)
        self.nom = tkinter.Label(self.frame, text="Nom du Client")
        self.nom.grid(column=1, row=0, padx=5, sticky="w")
        self.entre1= tkinter.Entry(self.frame, width =30)
        self.entre1.grid(column =2, row=0)
        
        #gereur despace entre les elements
        self.frame.grid_rowconfigure(1, minsize=50)
        #deuxieme element
        self.adresse = tkinter.Label(self.frame, text="Adresse/Quatier")
        self.adresse.grid(column=1, row=1, padx=5,sticky="w")
        self.entre2= tkinter.Entry(self.frame, width =30)
        self.entre2.grid(column =2, row=1)
        #troisieme element
        self.contact = tkinter.Label(self.frame, text="Contact")
        self.contact.grid(column=1, row=2, padx=5, sticky="w")
        self.entre3= tkinter.Entry(self.frame, width =30)
        self.entre3.config(validate="all", vcmd=(self.entre3.register(self.digit), '%S'))
        self.entre3.grid(column =2, row=2)
        
        # element und widgets controlle(informatoons sur l'achat)
        self.frame = tkinter.LabelFrame(self.tab1, text="Informations sur l'achat", padx=20, pady=10)
        self.frame.grid(column=0, row=1, padx=10, pady=10, sticky="w")
        self.nom = tkinter.Label(self.frame, text="Article")
        self.nom.grid(column=1, row=0, padx=5, sticky="w")
        self.choix= ttk.Combobox(self.frame,  values=[] )
        self.choix.insert(0,"Choisir un article")
        self.choix.config(foreground="grey")
        self.choix.bind("<<ComboboxSelected>>",self.handle_event)
        self.choix.bind('<FocusIn>', lambda event:self.focus_in(event, self.choix))
        self.choix.bind('<FocusOut>', lambda event:self.focus_out(event, self.choix))
        self.choix.grid(column =2, row=0)
        
        #gereur despace entre les elements
        self.frame.grid_rowconfigure(1, minsize=50)
        #deuxieme element
        self.date = tkinter.Label(self.frame, text="Date")
        self.date.grid(column=1, row=1, padx=5,sticky="w")
        self.showdate= DateEntry(self.frame, width =20, background='grey', foreground='white', locale='fr_FR',
                                 borderwidth=2, date_pattern='dd.MM.yyyy', selectbackground = "red", 
                                 normalbackground = "lightgreen",
                                 weekendbackground = "darkgreen",
                                 weekendforeground = "white")
        self.showdate.grid(column =2, row=1, padx=5)
        #troisieme element
        self.facture = tkinter.Label(self.frame, text="Facture.Nr")
        self.facture.grid(column=1, row=2, padx=5,sticky="w")
        self.showfacture= tkinter.Entry(self.frame, width =22)
        self.showfacture.grid(column =2, row=2)
        self.showfacture.insert(tk.END, self.generate_receipt_number())
        
        
        #gereur d'element deuxieme range sur les info sur l'article
        self.article = tkinter.Label(self.frame, text="Nombre d'article ")
        self.article.grid(column=3, row=0, padx=20, sticky="w")
        self.showarticle= tkinter.Spinbox(self.frame, width=18, from_= 1, to=100)
        self.showarticle.config(validate="all", vcmd=(self.showarticle.register(self.digit), '%S'))
        self.showarticle.grid(column =4, row=0)
        
        self.prix= tkinter.Label(self.frame, text="Prix Unitaire")
        self.prix.grid(column=3, row=1, padx=20, sticky="w")
        self.showprix= tkinter.Entry(self.frame, width=20)
        self.showprix.config(validate="all", vcmd=(self.showprix.register(self.digit), '%S'))
        self.showprix.bind("<KeyRelease>", lambda event: (self.summe(event), self.button_check(event)))
        self.showprix.grid(column =4, row=1)
        
        self.somme= tkinter.Label(self.frame, text="Somme")
        self.somme.grid(column=3, row=2, padx=20, sticky="w")
        self.showsomme= tkinter.Entry(self.frame, width=20)
        
        self.showsomme.grid(column =4, row=2)
        
        # gereur de 3 range
        
        self.nr = tkinter.Label(self.frame, text="d'Article.Nr")
        self.nr.grid(column=5, row=0, padx=10, sticky="e")
        self.shownr= tkinter.Entry(self.frame, width=20)
        self.shownr.grid(column =6, row=0)
        
        
        self.desc= tkinter.Label(self.frame, text="Description de  l'article")
        self.desc.grid(column=6, row=1, padx=20, sticky="w")
        self.showdesc= tk.Text(self.frame, width=20, height=3)
        self.showdesc.grid(column =6, row=2)
        self.placeholder_text = "Optionelle..."
        self.showdesc.insert("1.0", self.placeholder_text)
        self.showdesc.tag_configure("placeholder", foreground="grey")
        self.showdesc.tag_add("placeholder", "1.0", "end")
        self.showdesc.bind("<FocusIn>", self.remove_placeholder)
        self.showdesc.bind("<FocusOut>", self.add_placeholder)
        #liste des articles dans le panier
        style = ttk.Style()
        self.columns=('qty','model' ,'desc', 'price', 'total')
        self.treeview = ttk.Treeview(self.tab1, columns=self.columns, show='headings', height=8)
        self.treeview.heading("qty",text="Quantite")
        self.treeview.heading("model" , text="Model")
        self.treeview.heading("desc", text="Description")
        self.treeview.heading("price", text="Prix.Uni")
        self.treeview.heading("total", text="Somme")
        
        
        # Spaltenbreite anpassen
        self.treeview.column("qty", width=10)
        self.treeview.column("model",width=50)
        self.treeview.column("desc", width=140)
        self.treeview.column("price", width=100)
        self.treeview.column("total", width=50)
        self.treeview.grid(columnspan=4, column=0, row=4 ,rowspan=3 ,sticky='nsew', padx=10, pady=20)

        #tab2
        self.tab2 = ttk.Frame(self.tabControl)
        self.tabControl.add(self.tab2, text ='Espace de Recherche')
        self.tabControl.pack(expand = 1, fill ="both")
        
        # gestion avec les buttons
        self.Print = tkinter.Button(self.tab1, text="Vers Excel->", width=15, command=self.mettre_ds_excel)
        self.Print.place(relx=0.98, rely=0.95, anchor = tk.SE)
        self.Print.config(state="disabled", background="grey")
        
        #button pour pdf
        self.pdf = tkinter.Button(self.tab1, text="Facture", width=15, command=self.generate_invoice)
        self.pdf.place(relx=0.82, rely=0.95, anchor = tk.SE)
        self.pdf.config(state="disabled", background="grey")
        #self.pdf.config(state="disabled", background="#9ed676")
        #button pour tout efface apresune erreur
        self.add = tkinter.Button(self.tab1, text="Mettrer au panier", width=15, command=lambda: (self.add_item(), self.check_add()))
        self.add.place(relx=0.16, rely=0.95, anchor = tk.SE)
        self.add.config(state="disabled", background="grey")
        
        self.clear = tkinter.Button(self.tab1, text="Effacer du panier", width=15, command =lambda: (self.remove_pro(),self.check_add()))
        self.clear.place(relx=0.32, rely=0.95, anchor = tk.SE)
        self.clear.config(state="disabled", background="grey")
        
    def openHelp(self):
        mb.showinfo("Solution", "veillez contacter Donald ou Yannick pour plus de Details en ce qui concerne l' utilisation de ce Software et aussi en cas de Problem. Merci!")
    
    def remove_placeholder(self, event):
        current_text = self.showdesc.get("1.0", "end-1c")
        if current_text.strip() == self.placeholder_text:
            self.showdesc.delete("1.0", "end-1c")
            self.showdesc.tag_remove("placeholder", "1.0", "end")

    def add_placeholder(self, event):
        current_text = self.showdesc.get("1.0", "end-1c")
        if not current_text.strip():
            self.showdesc.insert("1.0", self.placeholder_text)
            self.showdesc.tag_add("placeholder", "1.0", "end")
            
    def button_check(self, event):
        nr= self.shownr.get()
        price= self.showprix.get()
        art=self.showarticle.get()
        if nr and price and  art:
            self.add.config(state="normal", background="lightblue")
        else:
            self.add.config(state="disabled", background="grey")

    def check_add(self):
        if self.treeview.get_children():
            self.clear.config(state="normal", background="lightblue")
            self.Print.config(state="normal", background="#9ed676")
            self.pdf.config(state="normal", background="#9ed676")
        else:
            self.clear.config(state="disabled", background="grey")
            self.Print.config(state="disabled", background="grey")
            self.pdf.config(state="disabled", background="grey")

    def article_facture_nr(self):
        try:
            self.shownr.delete(0, tk.END)
            self.shownr.insert(tk.END, self.generate_article_number())
        except Exception as e :
            mb.showwarning("Attention", "Veillez choisir un Produit..!Merci")
        
    def handle_event(self, event):
        self.choix.config(foreground="black")
        self.article_facture_nr()
        
    invoice_list = [] 
    def add_item(self):
        if not self.showprix.get() or not self.choix.get():
            mb.showwarning('Avertissement', 'Veuillez saisir le prix et choisir le produit..')
        else:
            self.qty = int(self.showarticle.get())
            self.art= self.choix.get()
            self.desc = self.showdesc.get("1.0", tk.END)
            self.price = float(self.showprix.get())
            self.som = self.showsomme.get()
            invoice_item = [self.qty, self.art, self.desc, self.price, self.som]
            self.treeview.insert('',0, values=invoice_item)
            self.clear_item()
            self.invoice_list.append(invoice_item)
        
    def clear_item(self):
        self.choix.delete(0,'end')
        self.showarticle.delete(0, tkinter.END)
        self.showarticle.insert(0, "1")
        self.showdesc.delete("1.0", tk.END)
        self.showprix.delete(0, tkinter.END)
        self.showsomme.delete(0, tkinter.END)

   
    def remove_pro(self):
        selected_items = self.treeview.selection()
        for item_id in selected_items:
            item_values = [self.treeview.item(item_id, "values")[i] for i in range(5)]
            self.treeview.delete(item_id)
            
            for invoice_item in self.invoice_list:
                if invoice_item[1] == item_values[1]:  
                    self.invoice_list.remove(invoice_item)

    def generate_invoice(self):
        name = self.entre1.get()
        phone = self.entre3.get()
        adresse = self.entre2.get()
        if not (name and phone and adresse):
            messagebox.showinfo("Error", "Veuillez saisir les Informations de la clientèle.")
        else:
            doc = DocxTemplate("facture.docx")
            facture = self.showfacture.get()
            date = datetime.datetime.now().date()
            heure = datetime.datetime.now().time().strftime("%H:%M:%S")
            article = self.shownr.get()
            subtotal = sum(item[3] for item in self.invoice_list) 
            salestax = 0.1
            total = subtotal * (1 - salestax)
            doc = DocxTemplate("facture.docx")
            name = self.entre1.get()
            phone = self.entre3.get()
            adresse = self.entre2.get()
            facture=self.showfacture.get()
            date= datetime.datetime.now().date()
            heure=datetime.datetime.now().time().strftime("%H:%M:%S")
            article=self.shownr.get()
            subtotal = sum(item[3] for item in self.invoice_list) 
            salestax = 0.1
            total = subtotal*(1-salestax)
            doc.render({
                    "name":name, 
                    "facture":facture,
                    "date":date,
                    "article":article, 
                    "heure":heure, 
                    "phone":phone,
                    "adresse":adresse,
                    "invoice_list": self.invoice_list,
                    "subtotal":subtotal,
                    "salestax":str(salestax*100)+"%",
                    "total":total})
            
            doc_name = "facture" + name + datetime.datetime.now().strftime("%Y-%m-%d-%H%M%S") + ".docx"
            doc.save(doc_name)
            pdf_name = doc_name.replace(".docx", ".pdf")
            self.convert_to_pdf(doc_name, pdf_name)
            self.treeview.delete(*self.treeview.get_children())
            self.invoice_list.clear()
            mb.showinfo("Success", "Votre Facture a ete creer avec succes! Merci")

    def convert_to_pdf(self, word_doc_name, pdf_name):
        word = comtypes.client.CreateObject("Word.Application")
        file_in= os.path.abspath(word_doc_name)
        file_out= os.path.abspath(pdf_name)
        word.Visible = False
        doc = word.Documents.Open(file_in)
        doc.SaveAs(file_out, FileFormat=17)  # 17 entspricht dem PDF-Format
        doc.Close()
        word.Quit()


    def generate_receipt_number(self):
        return uuid.uuid4().hex[:8].upper()
    def generate_article_number(self):
        # Generiere une UUID
        uuid_str = uuid.uuid4().hex
        #retire les signes - et transforme en chiffre
        article_number = ''.join(filter(str.isdigit, uuid_str))
        # limite la longeur a 6 chiffre
        article_number = article_number[:6]

        return article_number
    
    def summe(self, event):
        art = self.choix.get()
        prix = self.showprix.get()
        na = self.showarticle.get()
        
        if prix and na:
            try:
                summe = float(prix) * float(na)
                self.showsomme.delete(0, tk.END)
                self.showsomme.insert(0, summe)
            except ValueError:
                mb.showerror("Erreur", "Veillez entrer un nombre valide pour le prix ou la quantité! Merci")
        else:
            mb.showerror("Erreur", "Veillez entrer un prix unitaire ou une quantité! Merci")

    def digit(self, text):
        if text.isdigit():
            return True
        else:
            return False
        
    def write_into(self):
        self.nom1.delete(0, tk.END)
        self.nom1.insert(0, "Yann Electronic")
        self.adresse2.delete(0, tk.END)
        self.adresse2.insert(0, "PK8 Ndokoti Douala")
        if self.selected_seller.get()==1:
            self.entre.delete(0, tk.END)   
            self.entre.insert(0, "0178 / 1962229")
        elif self.selected_seller.get()==2:
            self.entre.delete(0, tk.END)   
            self.entre.insert(0, "696 69 14 57")
        else:
            self.entre.delete(0, tk.END)   
            self.entre.insert(0, "*** Numero inconnue ***")
    def focus_in(self, event, combobox):
        if combobox.get()=="Choisir un article":
            combobox.delete(0,tk.END)
            combobox.config(foreground="black")
    def focus_out(self,event, combobox):
        if not combobox.get():
            combobox.insert(0, "Choisir un article")
            combobox.config(foreground="gray")
    def filter(self, event, combobox, options):
        search= combobox.get()
        filtered_options=[option for option in options if isinstance(option, str) and option.lower().startswith(search.lower())]
        combobox.config(values=sorted(filtered_options))

    def combobox_values(self, category):
        df = pd.read_csv("modell.csv", sep=";", encoding="ISO-8859-1")
        values = df[category].dropna().tolist()
        return sorted(values)
        
    def update_combobox_values(self):
        if self.selected_numark.get() == "DJ Controller":
            self.choix['values'] = self.combobox_values("DJ Controller")            
        elif self.selected_numark.get() == "Player":
            self.choix['values'] = self.combobox_values("Player")
        elif self.selected_numark.get() == "Mixers":
            self.choix['values'] = self.combobox_values("Mixers")
        elif self.selected_numark.get() == "Platines Vinyles":
            self.choix['values'] = self.combobox_values("Platines vinyles")
        elif self.selected_numark.get() == "Cartouches":
            self.choix['values'] = self.combobox_values("Cartouches")
        elif self.selected_numark.get() == "Accessoires":
            self.choix['values'] = self.combobox_values("Accessoires")
        elif self.selected_pioneer.get() == "Lecteur DJ/tourne-disque":
            self.choix['values'] = self.combobox_values("Lecteur DJ/tourne-disque")
        elif self.selected_pioneer.get() == "DJ Mixers":
            self.choix['values'] = self.combobox_values("DJ Mixers")
        elif self.selected_pioneer.get() == "DJ Controllern":
            self.choix['values'] = self.combobox_values("Dj Controllern")
        elif self.selected_pioneer.get() == "All-in-one-DJ-Systeme":
            self.choix['values'] = self.combobox_values("All-in-one-DJ-Systeme")
        elif self.selected_pioneer.get() == "Haut-Parleurs de Controle":
            self.choix['values'] = self.combobox_values("Haut-parleurs de controle")
        elif self.selected_pioneer.get() == "Appareil a Effet DJ":
            self.choix['values'] = self.combobox_values("Appareil a Effets DJ")
        elif self.selected_pioneer.get() == "Ecouteurs":
            self.choix['values'] = self.combobox_values("Ecouteurs")
        elif self.selected_pioneer.get() == "Accesoires":
            self.choix['values'] = self.combobox_values("Accesoires")
        elif self.selected_pioneer.get() == "Baffle":
            self.choix['values'] = self.combobox_values("Baffle")
        self.choix.bind('<KeyRelease>', lambda event: self.filter(event, self.choix,self.choix["values"]))

    def mettre_ds_excel(self):
        try:
            wb = load_workbook("Facture.xlsx")
        except FileNotFoundError:
            wb = Workbook()
            messagebox.showerror("Erreur", "Le Produit n'a pas ete telecharge dans Excel. Veillez verifier vos Donnes sur le Software!")

        ws = wb.active
        ws['A1'] = "Article"
        ws['B1'] = "Date"
        ws['C1'] = "Facture_nr"
        ws['D1'] = "Article_nr"
        ws['E1'] = "Prix Unitaire"
        ws['F1'] = "Somme"
        # Trouve la prochaine ligne
        next_row = ws.max_row + 1

        # ecrire les Donnnes entrent  dans le formulaire sur l'excell
        ws[f"A{next_row}"] = self.choix.get()
        ws[f"B{next_row}"] = self.showdate.get()
        ws[f"C{next_row}"] = self.showfacture.get()
        ws[f"D{next_row}"] = self.shownr.get()
        ws[f"E{next_row}"] = self.showprix.get()
        ws[f"F{next_row}"] = self.showsomme.get()
        # Speichere die Excel-Datei
        wb.save("Facture.xlsx")
        messagebox.showinfo("Succes", "Le Produit et ses Details ont ete telecharge dans Excel. Merci!")

root = tkinter.Tk()
root.title("YannElectronic")

w = 810 
h = 650 


ws = root.winfo_screenwidth() # Largeur de la fenetre
hs = root.winfo_screenheight() # longeur de la fenetre

# Parametre pour le tableau (Software)
x = (ws/2) - (w/2)
y = (hs/2) - (h/2)
#logo = Image.open("images.png")
#photo = ImageTk.PhotoImage(logo)
#root.iconphoto(False, photo)
root.geometry('%dx%d+%d+%d' % (w, h, x, y))
app = Myfacture(root)
app.mainloop()  

